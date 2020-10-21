from openpyxl import Workbook
from openpyxl import load_workbook

# openpyxl python读写excel文件
# 读取

# 加载并读取excel文件
wb = load_workbook("books.xlsx")

# 读取sheet页
sheet = wb[wb.sheetnames[0]]

# 获取单元格的内容  第二行开始第三行结束，(2,4)不含4
for row in range(2, 4):
    for col in range(65, 68):   # A-D   65-68
        index = chr(col) + str(row)
        print(sheet[index].value, end="\t")
    print("")

# 关闭工作簿
wb.close()
